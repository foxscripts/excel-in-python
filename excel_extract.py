import xlrd
import openpyxl

# book2 = openpyxl.Workbook()
# book2.create_sheet('Sample')
# Acquire a sheet by its name
# aSheet = book2['Sample']

location = ("D:/Code/Python/python projects/basic-python/Book1.xlsx")
book = xlrd.open_workbook(location)
sheet = book.sheet_by_index(0) 

list = ["BRITISH", "INDIAN"]

for i in range(sheet.nrows):
	for j in range(sheet.ncols):
		for x in list:
			if sheet.cell_value(i, j) == x:
				print(sheet.row_values(i))
				# Writing to sheet
				# aSheet.cell(row=2, column=2).value ='e'

# book2.save('Sample.xlsx')
